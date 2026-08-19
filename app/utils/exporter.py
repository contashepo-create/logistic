# -*- coding: utf-8 -*-
"""
محرك التصدير والطباعة الموحّد لكل جداول وتقارير النظام:
  - تصدير Excel (openpyxl)
  - تصدير PDF (عبر QtPrintSupport — يدعم العربية بشكل كامل)
  - طباعة مباشرة بتنسيق احترافي (ترويسة الشركة + جدول منسق + تذييل)
"""
from __future__ import annotations

import os
import re
import sqlite3

from PySide6.QtCore import QMarginsF
from PySide6.QtGui import QPageLayout
from PySide6.QtGui import QFont, QPageSize, QTextDocument
from PySide6.QtPrintSupport import QPrintDialog, QPrinter
from PySide6.QtWidgets import (QApplication, QFileDialog, QMessageBox, QWidget)

from ..core import repo

NUM_RE = re.compile(r"^-?[\d,]+(\.\d+)?$")


# ---------------------------------------------------------------------------
# HTML (يُستخدم للطباعة وتصدير PDF)
# ---------------------------------------------------------------------------
def _esc(v) -> str:
    return (str(v) if v is not None else "—").replace("&", "&amp;") \
        .replace("<", "&lt;").replace(">", "&gt;")


def company_header_html(conn: sqlite3.Connection) -> str:
    info = repo.company_info(conn)
    parts = [f"<b style='font-size:16pt'>{_esc(info['company_name'])}</b>"]
    contact = " | ".join(x for x in (info.get("company_phone"),
                                     info.get("company_address")) if x)
    if contact:
        parts.append(contact)
    return "<br>".join(parts)


def build_table_html(headers: list[str], rows: list[list],
                     center_from: int | None = None) -> str:
    """جدول RTL منسّق. center_from: من هذا العمود فصاعداً توسيط (أعمدة الأرقام)."""
    out = ["<table dir='rtl' width='100%' border='0.5' cellspacing='0' "
           "cellpadding='4' style='font-size:10pt'>",
           "<tr bgcolor='#1f4e79'>"]
    for h in headers:
        out.append(f"<td align='center'><b><font color='white'>{_esc(h)}</font></b></td>")
    out.append("</tr>")
    for i, row in enumerate(rows):
        bg = "#f2f6fa" if i % 2 else "white"
        out.append(f"<tr bgcolor='{bg}'>")
        for c, cell in enumerate(row):
            if center_from is not None and c >= center_from:
                align = "center"
            else:
                align = "right"
            out.append(f"<td align='{align}'>{_esc(cell)}</td>")
        out.append("</tr>")
    out.append("</table>")
    return "".join(out)


def build_report_html(conn: sqlite3.Connection, title: str, subtitle: str = "",
                      headers: list[str] | None = None,
                      rows: list[list] | None = None,
                      summary_lines: list[tuple[str, str]] | None = None,
                      center_from: int | None = None,
                      footer_note: str = "") -> str:
    """تقرير احترافي: ترويسة الشركة + عنوان + جدول + ملخصات + تذييل."""
    body = [f"<div align='center'>{company_header_html(conn)}</div><hr>"]
    body.append(f"<div align='center' style='font-size:14pt'><b>{_esc(title)}</b></div>")
    if subtitle:
        body.append(f"<div align='center'>{_esc(subtitle)}</div><br>")
    if summary_lines:
        body.append("<table dir='rtl' width='60%' align='center' border='0.5' "
                    "cellspacing='0' cellpadding='3' style='font-size:10pt'>")
        for k, v in summary_lines:
            body.append(
                f"<tr><td align='right'>{_esc(k)}</td>"
                f"<td align='center'><b>{_esc(v)}</b></td></tr>")
        body.append("</table><br>")
    if headers:
        body.append(build_table_html(headers, rows or [], center_from=center_from))
        body.append(f"<div align='left' style='font-size:9pt'>"
                    f"عدد السجلات: {len(rows or [])}</div>")
    note = footer_note or repo.get_setting(conn, "company_vat_note", "")
    if note:
        body.append(f"<br><div align='center' style='font-size:9pt'>{_esc(note)}</div>")
    return "".join(body)


# ---------------------------------------------------------------------------
# حوار حفظ الملف
# ---------------------------------------------------------------------------
def _safe_filename(name: str) -> str:
    """تطهير اسم الملف من محارف غير صالحة (مسارات فاصلة إلخ)."""
    import re as _re
    clean = _re.sub(r'[\\/:*?"<>|]', "-", str(name)).strip().strip(".")
    return clean[:120] or "export"


def _ask_save_path(parent: QWidget, default_name: str, patterns: list[str]) -> str | None:
    default_name = _safe_filename(default_name)
    path, _ = QFileDialog.getSaveFileName(parent, "حفظ الملف", default_name,
                                          ";;".join(patterns))
    return path or None


def _ask_overwrite(parent: QWidget, path: str) -> bool:
    if os.path.exists(path):
        ret = QMessageBox.question(
            parent, "الملف موجود",
            f"الملف التالي موجود مسبقاً:\n{path}\nهل تريد استبداله؟")
        return ret == QMessageBox.StandardButton.Yes
    return True


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def _notify(parent: QWidget | None, text: str) -> None:
    """إظهار رسالة نجاح (تُتجاهل في وضع الاختبار الآلي بدون نوافذ)."""
    import os
    if os.environ.get("LOGISTIC_HEADLESS"):
        return
    QMessageBox.information(parent, "تم", text)


def export_excel(parent: QWidget, conn: sqlite3.Connection, title: str,
                 headers: list[str], rows: list[list], default_name: str,
                 summary_lines: list[tuple[str, str]] | None = None) -> str | None:
    """تصدير جدول إلى Excel مع ورقة RTL وترويسة منسقة."""
    default_name = _safe_filename(default_name)
    path = _ask_save_path(parent, default_name, ["Excel (*.xlsx)"])
    if not path:
        return None
    if not _ask_overwrite(parent, path):
        return None
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير"
        ws.sheet_view.rightToLeft = True

        info = repo.company_info(conn)
        n_cols = max(len(headers), 2)
        thin = Side(style="thin", color="B0B0B0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        c = ws.cell(row=1, column=1, value=info["company_name"])
        c.font = Font(bold=True, size=16, color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        c = ws.cell(row=2, column=1, value=title)
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="center")
        r = 3
        if summary_lines:
            for k, v in summary_lines:
                ws.cell(row=r, column=1, value=k).font = Font(bold=True)
                cell = ws.cell(row=r, column=2, value=v)
                cell.alignment = Alignment(horizontal="center")
                r += 1
            r += 1
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
        ws.row_dimensions[r].height = 22
        header_row = r
        r += 1
        for row in rows:
            for i, v in enumerate(row, start=1):
                s = str(v) if v is not None else ""
                cell = ws.cell(row=r, column=i)
                if NUM_RE.match(s.replace("٫", ".")):
                    cell.value = float(s.replace(",", ""))
                    cell.number_format = "#,##0.00"
                else:
                    cell.value = s
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            r += 1
        # عرض الأعمدة تقريبي حسب أطول قيمة
        for i, h in enumerate(headers, start=1):
            width = len(str(h))
            for row in rows:
                if i <= len(row):
                    width = max(width, len(str(row[i - 1])))
            ws.column_dimensions[get_column_letter(i)].width = min(max(width + 4, 10), 40)
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        wb.save(path)
        _notify(parent, f"تم تصدير الملف بنجاح:\n{path}")
        return path
    except Exception as e:  # noqa: BLE001
        if os.environ.get("LOGISTIC_HEADLESS"):
            raise
        QMessageBox.critical(parent, "خطأ في التصدير", str(e))
        return None


# ---------------------------------------------------------------------------
# PDF + الطباعة
# ---------------------------------------------------------------------------
def _make_doc(html: str) -> QTextDocument:
    from .fonts import pick_font_family
    doc = QTextDocument()
    doc.setDefaultFont(QFont(pick_font_family(), 10))
    doc.setHtml(html)
    return doc


def _pdf_to_file(html: str, path: str, printer: QPrinter | None = None) -> None:
    """كتابة مستند PDF إلى مسار (مشترك بين تصدير PDF والطباعة)."""
    pr = printer
    if pr is None:
        pr = QPrinter(QPrinter.PrinterMode.HighResolution)
        pr.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        pr.setOutputFileName(path)
        pr.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
    pr.setPageMargins(QMarginsF(12, 12, 12, 14), QPageLayout.Unit.Millimeter)
    doc = _make_doc(html)
    doc.print_(pr)


def export_pdf(parent: QWidget, html: str, default_name: str) -> str | None:
    default_name = _safe_filename(default_name)
    path = _ask_save_path(parent, default_name, ["PDF (*.pdf)"])
    if not path:
        return None
    if not _ask_overwrite(parent, path):
        return None
    if not path.lower().endswith(".pdf"):
        path += ".pdf"
    try:
        _pdf_to_file(html, path)
        _notify(parent, f"تم إنشاء ملف PDF:\n{path}")
        return path
    except Exception as e:  # noqa: BLE001
        if os.environ.get("LOGISTIC_HEADLESS"):
            raise
        QMessageBox.critical(parent, "خطأ في تصدير PDF", str(e))
        return None


def print_html(parent: QWidget, html: str) -> None:
    """فتح حوار الطباعة وطباعة المستند بتنسيق احترافي.

    في وضع الاختبار الآلي (LOGISTIC_HEADLESS) يُولَّد ملف PDF مؤقت
    بدلاً من فتح حوار الطباعة.
    """
    if os.environ.get("LOGISTIC_HEADLESS"):
        import tempfile
        path = os.path.join(tempfile.gettempdir(), "logistic_print_test.pdf")
        _pdf_to_file(html, path)
        return
    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
    dlg = QPrintDialog(printer, parent)
    dlg.setWindowTitle("طباعة")
    if dlg.exec() == QDialog.DialogCode.Accepted:
        doc = _make_doc(html)
        doc.print_(printer)


from PySide6.QtWidgets import QDialog  # noqa: E402  (مطلوب بعد التعريفات أعلاه)
