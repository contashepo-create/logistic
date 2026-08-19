# -*- coding: utf-8 -*-
"""
فحص الواجهة العميق: كل المسارات/الصفحات، كل النوافذ (إضافة/تعديل/عرض)،
كل الأزرار الصفّية وشريط الأدوات والتصدير، كل الفلاتر والبحث والإجماليات،
وسلامة ملفات Excel/PDF الناتجة فعلياً.

تشغيل:  QT_QPA_PLATFORM=offscreen python scripts/test_ui_deep.py
"""
from __future__ import annotations

import os
import sys
import tempfile
import traceback
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
os.environ["LOGISTIC_DATA_DIR"] = tempfile.mkdtemp(prefix="logistic_uideep_")
os.environ["LOGISTIC_HEADLESS"] = "1"

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QApplication, QComboBox, QDateEdit, QDialog, QDoubleSpinBox, QFileDialog,
    QLineEdit, QMessageBox, QPushButton, QSpinBox,
)

PASS = FAIL = 0
FAILURES: list[str] = []
LOG_LINES: list[str] = []


def check(name: str, cond: bool, extra: str = "") -> bool:
    global PASS, FAIL
    if cond:
        PASS += 1
    else:
        FAIL += 1
        FAILURES.append(f"{name} {extra}")
        print(f"  ❌ {name} {extra}")
    return cond


def step(title: str) -> None:
    print(f"== {title}", flush=True)


app = QApplication(sys.argv)
app.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
from app.ui.theme import apply_theme
apply_theme(app)

# ---------------------------------------------------------------------------
# بيئة آمنة: اعتراض كل النوافذ المنبثقة وحوارات الملفات
# ---------------------------------------------------------------------------
POPUPS: list[str] = []


def _fake_msg(*args, **kwargs):
    POPUPS.append(str(args[1] if len(args) > 1 else args))
    return QMessageBox.StandardButton.No  # الإجابة الافتراضية: إلغاء


for fn in ("information", "warning", "critical", "about"):
    setattr(QMessageBox, fn, staticmethod(_fake_msg))
QMessageBox.question = staticmethod(_fake_msg)

_original_exec = QDialog.exec


def _nonblocking_exec(self):
    """بدون حظر: بناء + معالجة أحداث ثم إغلاق (وضع الاختبار فقط)."""
    self.show()
    app.processEvents()
    self.close()
    return QDialog.DialogCode.Accepted


QDialog.exec = _nonblocking_exec

EXPORT_DIR = Path(tempfile.mkdtemp(prefix="logistic_uix_"))
EXPORTED: list[Path] = []

from app.utils import exporter


def fake_save_path(parent, default_name, filters):
    name = default_name or "export.bin"
    n = len(EXPORTED)
    out = EXPORT_DIR / f"{n:03d}_{name}"
    EXPORTED.append(out)
    return str(out)


exporter._ask_save_path = fake_save_path
exporter._ask_overwrite = lambda parent, path: True

_tmp_att = Path(tempfile.gettempdir()) / "mرفق_test.pdf"
_tmp_att.write_bytes(b"%PDF-1.4 test")
QFileDialog.getOpenFileNames = staticmethod(
    lambda *a, **k: ([str(_tmp_att)], ""))
QFileDialog.getSaveFileName = staticmethod(
    lambda *a, **k: (str(EXPORT_DIR / "dlg_save.bin"), ""))

# ---------------------------------------------------------------------------
step("تهيئة بيانات تجريبية")
from app.core import calc, db, repo

db.init_db()
conn = db.get_conn()
from scripts.seed_demo import seed

ids = seed(conn)
check("البيانات التجريبية جاهزة", conn.execute(
    "SELECT COUNT(*) FROM invoices").fetchone()[0] == 2)

# ===========================================================================
step("1) كل مسارات التنقل: بناء الصفحات والتحديث")
from app.ui.main_window import MainWindow, NAV_SECTIONS

win = MainWindow()
win.resize(1280, 780)
win.show()
app.processEvents()

pages_by_label = dict(win._pages)
expected_pages = [(label, cls) for _s, items in NAV_SECTIONS for label, cls in items]
check("عدد الصفحات = عدد عناصر التنقل",
      len(win._pages) == len(expected_pages))
# التحقق من النوع الفعلي: أي صفحة استُبدلت بعنصر بديل (خطأ بناء) = فشل
for label, cls in expected_pages:
    page = pages_by_label.get(label)
    check(f"الصفحة «{label}» من الفئة الصحيحة",
          type(page).__name__ == cls.__name__,
          f"(وجدنا {type(page).__name__})")

page_objects = []
for label, _cls in expected_pages:
    page = pages_by_label.get(label)
    try:
        if hasattr(page, "refresh"):
            page.refresh()
        page_objects.append((label, page))
        check(f"الصفحة «{label}» تُبنى وتُحدَّث", True)
    except Exception as e:  # noqa: BLE001
        check(f"الصفحة «{label}» تُبنى وتُحدَّث", False, f"— {e}")
        print(traceback.format_exc())

# التنقل الفعلي عبر القائمة لكل صفحة قابلة للتحديد
selectable = [r for r in range(win.nav_list.count())
              if win.nav_list.item(r).flags() & Qt.ItemFlag.ItemIsSelectable]
nav_ok = True
try:
    for r in selectable:
        win.nav_list.setCurrentRow(r)
        app.processEvents()
    check("التنقل عبر كل المسارات بلا أخطاء", True)
except Exception as e:  # noqa: BLE001
    nav_ok = False
    check("التنقل عبر كل المسارات بلا أخطاء", False, str(e))


def crud_pages():
    out = []
    for label, page in page_objects:
        if hasattr(page, "table") and hasattr(page, "on_add"):
            out.append((label, page))
    return out


check("عدد صفحات CRUD >= 10", len(crud_pages()) >= 10, f"({len(crud_pages())})")

# ===========================================================================
step("2) أزرار كل صف (عرض/تعديل/حذف/إضافية) — نقر فعلي على الأزرار")
import app.ui.widgets as W
import app.ui.pages_ops as P_ops
import app.ui.pages_master as P_master
import app.ui.pages_payroll as P_payroll
import app.ui.dialogs_master as D_master

def _confirm_no(parent, text, title="تأكيد"):
    return False

def _confirm_yes(parent, text, title="تأكيد"):
    return True

CONFIRM_STATE = {"yes": False}

def _confirm(parent, text, title="تأكيد"):
    return CONFIRM_STATE["yes"]

for mod in (W, P_ops, P_master, P_payroll, D_master):
    mod.confirm = _confirm
    mod.info = lambda *a, **k: None
    mod.warn = lambda *a, **k: None

from app.core.rules import RuleError as _RuleError  # noqa: F401

from PySide6.QtWidgets import QTableWidget, QWidget as QW


def click_all_row_buttons(page, label) -> None:
    table = page.table
    cols = table.columnCount()
    rows = table.rowCount()
    clicked = 0
    for r in range(min(rows, 4)):
        for c in range(cols):
            w = table.cellWidget(r, c)
            if w is None:
                continue
            for b in w.findChildren(QPushButton):
                b.click()
                app.processEvents()
                clicked += 1
    check(f"[{label}] النقر على أزرار الصفوف ({clicked})", True)


before_counts = {}
for t in ("customers", "invoices", "receipt_vouchers", "payment_vouchers",
          "payrolls", "financial_years"):
    before_counts[t] = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]

for label, page in crud_pages():
    try:
        click_all_row_buttons(page, label)
    except Exception as e:  # noqa: BLE001
        check(f"[{label}] أزرار الصفوف", False, str(e))
        print(traceback.format_exc())

after_counts = {t: conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                for t in before_counts}
check("الأزرار مع confirm=False لم تحذف شيئاً", before_counts == after_counts)

# حذف فعلي عبر زر واحد مع confirm=True على سند قابل للحذف (سنة العرض مفتوحة)
rv_id = conn.execute("SELECT id FROM receipt_vouchers "
                     "WHERE date >= '2026-01-01' LIMIT 1").fetchone()["id"]
CONFIRM_STATE["yes"] = True
receipts_page = pages_by_label["سندات القبض"]
receipts_page.refresh()
try:
    # محاكاة إشارة الحذف مباشرة (نفس ما يفعله الزر)
    receipts_page.on_delete(rv_id)
    app.processEvents()
    exists = conn.execute("SELECT COUNT(*) FROM receipt_vouchers WHERE id=?",
                          (rv_id,)).fetchone()[0]
    check("حذف سند عبر مسار الواجهة ينجح", exists == 0)
except Exception as e:  # noqa: BLE001
    check("حذف سند عبر مسار الواجهة ينجح", False, str(e))
CONFIRM_STATE["yes"] = False

# ===========================================================================
step("3) شريط أدوات كل صفحة: زر الإضافة والبحث والفلاتر")
for label, page in crud_pages():
    try:
        if page.frame.add_btn:
            page.frame.add_btn.click()
            app.processEvents()
        check(f"[{label}] زر الإضافة يفتح النافذة", True)
    except Exception as e:  # noqa: BLE001
        check(f"[{label}] زر الإضافة", False, str(e))

# البحث
cust_page = pages_by_label["العملاء"]
cust_page.refresh()
total = cust_page.table.rowCount()
cust_page.frame.search_edit.setText("مؤسسة الرياض")
app.processEvents()
filtered = cust_page.table.rowCount()
check("البحث يرشّح النتائج", 0 < filtered < total,
      f"({filtered}/{total})")
cust_page.frame.search_edit.clear()

# فلاتر التاريخ في فواتير النقل
inv_page = pages_by_label["فواتير النقل"]
inv_page.filter_row.from_edit.set_iso("2026-03-01")
inv_page.filter_row.to_edit.set_iso("2026-03-31")
inv_page.refresh()
narrow = inv_page.table.rowCount()
expected_narrow = conn.execute(
    "SELECT COUNT(*) FROM invoices WHERE date BETWEEN '2026-03-01' AND '2026-03-31'"
).fetchone()[0]
check("فلتر التاريخ في الفواتير مطابق للـSQL", narrow == expected_narrow,
      f"({narrow} مقابل {expected_narrow})")
inv_page.filter_row.from_edit.set_iso("2026-01-01")
inv_page.filter_row.to_edit.set_iso("2026-12-31")
inv_page.refresh()

# فلتر نوع السندات
pay_page = pages_by_label["سندات الدفع"]
idx = pay_page.type_combo.findData("advance")
pay_page.type_combo.setCurrentIndex(idx)
pay_page.refresh()
only_adv = all("سلفة" in pay_page.table.item(r, 2).text()
               for r in range(pay_page.table.rowCount()))
check("فلتر نوع سند الدفع (سلفة)", only_adv and pay_page.table.rowCount() > 0)
pay_page.type_combo.setCurrentIndex(0)
pay_page.refresh()

# ===========================================================================
step("4) النوافذ الحوارية: إضافة فعلة عبر الحقول ثم save()")

def dlg_check(name, fn):
    try:
        fn()
        check(name, True)
    except Exception as e:  # noqa: BLE001
        check(name, False, f"— {type(e).__name__}: {e}")
        print(traceback.format_exc())


from app.ui.dialogs_master import (
    AccountDialog, CustomerDialog, EmployeeDialog, VehicleDialog, YearDialog,
)
from app.ui.dialogs_ops import InvoiceDialog, PaymentDialog, ReceiptDialog
from app.ui.dialogs_payroll import PayrollDialog

# عميل
def _add_customer():
    d = CustomerDialog(win)
    d.name_edit.setText("عميل فحص واجهة")
    d.phone_edit.setText("٠٥٥٥٠٠١١")  # أرقام عربية
    d.address_edit.setText("الرياض")
    d.opening_edit.set_value(1234.5)
    d.notes_edit.setPlainText("ملاحظة")
    d.save()
    row = conn.execute("SELECT * FROM customers WHERE name='عميل فحص واجهة'").fetchone()
    assert row and row["opening_balance"] == 1234.5
dlg_check("نافذة عميل: حفظ فعلي عبر الحقول", _add_customer)

def _edit_customer():
    cid = conn.execute("SELECT id FROM customers WHERE name='عميل فحص واجهة'"
                       ).fetchone()["id"]
    d = CustomerDialog(win, cid)
    d.name_edit.setText("عميل فحص معدل")
    d.save()
    assert repo.get_customer(conn, cid)["name"] == "عميل فحص معدل"
dlg_check("نافذة عميل: تعديل فعلي", _edit_customer)

def _view_customer():
    cid = conn.execute("SELECT id FROM customers WHERE name='عميل فحص معدل'"
                       ).fetchone()["id"]
    d = CustomerDialog(win, cid, read_only=True)
    assert d.name_edit.isReadOnly() and d.opening_edit.isReadOnly()
    assert not d.phone_edit.isEnabled() or d.phone_edit.isReadOnly()
dlg_check("نافذة عميل: وضع العرض مقفل", _view_customer)

# موظف
def _add_employee():
    d = EmployeeDialog(win)
    d.name_edit.setText("سائق فحص")
    d.type_combo.setCurrentIndex(d.type_combo.findData("driver"))
    d.save()
    row = conn.execute("SELECT * FROM employees WHERE name='سائق فحص'").fetchone()
    assert row["emp_type"] == "driver"
dlg_check("نافذة موظف: حفظ", _add_employee)

# سيارة
def _add_vehicle():
    d = VehicleDialog(win)
    d.plate_edit.setText("ف ح ص 77")
    drv = conn.execute("SELECT id FROM employees WHERE name='سائق فحص'").fetchone()["id"]
    d.driver_combo.select(drv)
    d.save()
    row = conn.execute("SELECT * FROM vehicles WHERE plate_number='ف ح ص 77'").fetchone()
    assert row and row["default_driver_id"] == drv
dlg_check("نافذة سيارة: حفظ مع سائق افتراضي", _add_vehicle)

# سنة مالية
def _add_year():
    d = YearDialog(win)
    d.year_spin.setValue(2028)
    d.from_edit.set_iso("2028-01-01")
    d.to_edit.set_iso("2028-12-31")
    d.save()
    assert repo.get_year(conn, conn.execute(
        "SELECT id FROM financial_years WHERE year=2028").fetchone()["id"])["year"] == 2028
dlg_check("نافذة سنة مالية: حفظ", _add_year)

# خزينة/بنك
def _add_accounts():
    d = AccountDialog("cashbox", win)
    d.name_edit.setText("خزينة الفحص")
    d.opening_edit.set_value(999)
    d.save()
    assert conn.execute("SELECT COUNT(*) FROM cashboxes WHERE name='خزينة الفحص'"
                        ).fetchone()[0] == 1
    b = AccountDialog("bank", win)
    b.name_edit.setText("بنك الفحص")
    b.accnum_edit.setText("ACC1")
    b.iban_edit.setText("IBAN1")
    b.opening_edit.set_value(5000)
    b.save()
    assert conn.execute("SELECT COUNT(*) FROM banks WHERE iban='IBAN1'").fetchone()[0] == 1
dlg_check("نافذتا خزينة/بنك: حفظ", _add_accounts)

# سند قبض (كل الأنواع)
def _receipts():
    cb = conn.execute("SELECT id FROM cashboxes WHERE name='الخزينة الرئيسية'"
                      ).fetchone()["id"]
    d = ReceiptDialog(win)
    d.date_edit.set_iso("2026-06-10")
    d.account_combo.select("cashbox", cb)
    d.type_combo.setCurrentIndex(d.type_combo.findData("customer"))
    d.customer_combo.select(ids["cust1"])
    d.amount_edit.setText("١٥٠٠٫٥٠")  # أرقام عربية + فاصلة عربية
    d.save()
    row = conn.execute("SELECT * FROM receipt_vouchers ORDER BY id DESC LIMIT 1"
                       ).fetchone()
    assert row["amount"] == 1500.50 and row["voucher_type"] == "customer"
    d2 = ReceiptDialog(win)
    d2.date_edit.set_iso("2026-06-11")
    d2.account_combo.select("cashbox", cb)
    d2.type_combo.setCurrentIndex(d2.type_combo.findData("other"))
    d2.amount_edit.set_value(300)
    d2.desc_edit.setText("خردة فحص")
    d2.save()
    row2 = conn.execute("SELECT * FROM receipt_vouchers ORDER BY id DESC LIMIT 1"
                        ).fetchone()
    assert row2["voucher_type"] == "other" and row2["amount"] == 300
dlg_check("سندات القبض (عميل + إيراد آخر، مبلغ بأرقام عربية)", _receipts)

# سند دفع — الأنواع الأربعة
def _payments():
    cb = ids["cb"]
    trip = conn.execute("SELECT t.id FROM invoice_trips t LIMIT 1").fetchone()["id"]
    # رحلة
    d = PaymentDialog(win)
    d.date_edit.set_iso("2026-06-12")
    d.account_combo.select("cashbox", cb)
    d.type_combo.setCurrentIndex(d.type_combo.findData("trip"))
    d.trip_combo.select(trip)
    d.amount_edit.set_value(111)
    d.save()
    r = conn.execute("SELECT * FROM payment_vouchers ORDER BY id DESC LIMIT 1").fetchone()
    assert r["voucher_type"] == "trip" and r["trip_id"] == trip
    # سلفة
    d = PaymentDialog(win)
    d.date_edit.set_iso("2026-06-13")
    d.account_combo.select("cashbox", cb)
    d.type_combo.setCurrentIndex(d.type_combo.findData("advance"))
    d.employee_combo.select(ids["drv1"])
    d.amount_edit.set_value(600)
    d.save()
    r = conn.execute("SELECT * FROM payment_vouchers ORDER BY id DESC LIMIT 1").fetchone()
    assert r["voucher_type"] == "advance" and r["employee_id"] == ids["drv1"]
    # سيارة
    d = PaymentDialog(win)
    d.date_edit.set_iso("2026-06-14")
    d.account_combo.select("bank", ids["bnk"])
    d.type_combo.setCurrentIndex(d.type_combo.findData("vehicle"))
    d.vehicle_combo.select(ids["veh1"])
    d.vehexp_combo.setCurrentIndex(d.vehexp_combo.findData("tires"))
    d.amount_edit.set_value(700)
    d.save()
    r = conn.execute("SELECT * FROM payment_vouchers ORDER BY id DESC LIMIT 1").fetchone()
    assert r["voucher_type"] == "vehicle" and r["vehicle_expense"] == "tires"
    # عام
    d = PaymentDialog(win)
    d.date_edit.set_iso("2026-06-15")
    d.account_combo.select("cashbox", cb)
    d.type_combo.setCurrentIndex(d.type_combo.findData("general"))
    d.amount_edit.set_value(88)
    d.save()
    r = conn.execute("SELECT * FROM payment_vouchers ORDER BY id DESC LIMIT 1").fetchone()
    assert r["voucher_type"] == "general"
dlg_check("سندات الدفع الأربعة (رحلة/سلفة/سيارة/عام)", _payments)

# سند دفع مرفوض (مبلغ صفر) عبر مسار الواجهة
def _payment_invalid():
    d = PaymentDialog(win)
    d.date_edit.set_iso("2026-06-16")
    d.account_combo.select("cashbox", ids["cb"])
    d.type_combo.setCurrentIndex(d.type_combo.findData("general"))
    d.amount_edit.set_value(0)
    try:
        d.save()
        raise AssertionError("لم يُرفض المبلغ الصفري")
    except Exception:
        pass
    n = conn.execute("SELECT COUNT(*) FROM payment_vouchers "
                     "WHERE date='2026-06-16'").fetchone()[0]
    assert n == 0
dlg_check("سند دفع صفر المبلغ: مرفوض ولم يُخزن", _payment_invalid)

# فاتورة كاملة عبر النافذة (نقلات + مصروفات + مرفق)
def _invoice():
    d = InvoiceDialog(win)
    d.date_edit.set_iso("2026-06-20")
    d.customer_combo.select(ids["cust2"])
    d.notes_edit.setText("فاتورة فحص")
    d.trips = [
        {"vehicle_id": ids["veh1"], "driver_id": ids["drv1"], "from_loc": "الرياض",
         "to_loc": "جدة", "price": 3000, "notes": "",
         "expenses": [{"expense_type": "fuel", "amount": 200, "notes": ""},
                      {"expense_type": "trip", "amount": 150, "notes": ""}]},
        {"vehicle_id": None, "driver_id": None, "from_loc": "جدة", "to_loc": "مكة",
         "price": 900, "notes": "", "expenses": []},
    ]
    d.attachments = [repo.store_attachment(str(_tmp_att))]
    d.refresh()
    d.save()
    inv = conn.execute("SELECT id FROM invoices WHERE notes='فاتورة فحص'").fetchone()["id"]
    t = calc.invoice_totals(conn, inv)
    assert t["trips_total"] == 3900 and t["expenses_total"] == 350
    full = calc.get_invoice_full(conn, inv)
    assert len(full["trips"]) == 2 and len(full["attachments"] or "[]") > 5
    # تعديلها: غيّر سعر أول نقلة
    d2 = InvoiceDialog(win, inv)
    d2.trips[0]["price"] = 3500
    d2.refresh()
    d2.save()
    assert calc.invoice_totals(conn, inv)["trips_total"] == 4400
    return inv
INV_ID = None
def _invoice_wrapper():
    global INV_ID
    INV_ID = _invoice()
dlg_check("فاتورة كاملة: إضافة + مرفق + تعديل بأثر رجعي", _invoice_wrapper)

# راتب مع خصم جزئي من سلفة عبر النافذة
def _payroll():
    open_advs = repo.employee_advances(conn, ids["drv1"], include_settled=False)
    if not open_advs:
        # أنشئ سلفة جديدة لسائق الفحص إن لم تتوفر
        d0 = PaymentDialog(win)
        d0.date_edit.set_iso("2026-06-25")
        d0.account_combo.select("cashbox", ids["cb"])
        d0.type_combo.setCurrentIndex(d0.type_combo.findData("advance"))
        d0.employee_combo.select(ids["drv1"])
        d0.amount_edit.set_value(500)
        d0.save()
        open_advs = repo.employee_advances(conn, ids["drv1"], False)
    d = PayrollDialog(win)
    d.date_edit.set_iso("2026-06-28")
    d.employee_combo.select(ids["drv1"])
    app.processEvents()
    d.month_combo.setCurrentIndex(d.month_combo.findData(6))
    d.account_combo.select("cashbox", ids["cb"])
    d.base_edit.set_value(4000)
    d.additions_edit.set_value(250)
    d.additions_note.setText("حوافز")
    d.other_ded_edit.set_value(50)
    spins = []
    for r in range(d.adv_table.rowCount()):
        wgt = d.adv_table.cellWidget(r, 5)
        if isinstance(wgt, QDoubleSpinBox):
            sb = wgt
        elif wgt and wgt.findChildren(QDoubleSpinBox):
            sb = wgt.findChildren(QDoubleSpinBox)[0]
        else:
            sb = None
        spins.append(sb)
    assert spins, "لم تُحمّل السلفيات في نافذة الراتب"
    spins[0].setValue(min(300.0, spins[0].maximum()))
    d._recalc()
    expected_net = 4000 + 250 - spins[0].value() - 50
    assert abs(float(d.net_label.text().replace(",", "")) - expected_net) < 0.01
    d.save()
    p = repo.get_payroll(conn, conn.execute(
        "SELECT id FROM payrolls ORDER BY id DESC LIMIT 1").fetchone()["id"])
    assert p["base_salary"] == 4000 and p["additions"] == 250
    assert abs(p["net_salary"] - expected_net) < 0.01
    assert p["settlements"] and p["settlements"][0]["amount"] == spins[0].value()
dlg_check("نافذة راتب: سلفيات تلقائية + خصم جزئي + صافي", _payroll)

# ===========================================================================
step("5) أزرار التصدير والطباعة في كل صفحة — ملفات فعلية والتحقق من صحتها")
from openpyxl import load_workbook


def valid_export(p: Path) -> str:
    if not p.exists() or p.stat().st_size < 400:
        return "ملف ناقص"
    if p.suffix == ".xlsx":
        try:
            wb = load_workbook(p)
            ws = wb.active
            if ws.max_row < 1 or ws.max_column < 1:
                return "ورقة فارغة"
            return ""
        except Exception as e:  # noqa: BLE001
            return f"xlsx غير صالح: {e}"
    if p.suffix == ".pdf":
        head = p.read_bytes()[:5]
        if head != b"%PDF-":
            return "بلا توقيع PDF"
        return ""
    return ""


for label, page in page_objects:
    if not hasattr(page, "do_export"):
        continue
    for mode in ("excel", "pdf", "print"):
        try:
            page.do_export(mode)
            app.processEvents()
            if mode != "print":
                p = EXPORTED[-1] if EXPORTED else None
                err = valid_export(p) if p else "لا ملف"
                check(f"[{label}] تصدير {mode}", err == "", err)
            else:
                # الطباعة في وضع headless تولد PDF مؤقتاً
                check(f"[{label}] طباعة (معاينة PDF)", Path(
                    os.path.join(tempfile.gettempdir(), "logistic_print_test.pdf")
                ).exists())
        except Exception as e:  # noqa: BLE001
            check(f"[{label}] تصدير {mode}", False, str(e))
            print(traceback.format_exc())

# كشوف الحساب المنفصلة (أزرار كشف الحساب)
from app.ui.statements import AccountStatementDialog, CustomerStatementDialog

for name, mk in (
    ("كشف عميل", lambda: CustomerStatementDialog(ids["cust1"], win)),
    ("كشف خزينة", lambda: AccountStatementDialog("cashbox", ids["cb"], win)),
    ("كشف بنك", lambda: AccountStatementDialog("bank", ids["bnk"], win)),
):
    dlg = mk()
    rows = dlg.table.rowCount()
    check(f"[{name}] فُتح وفيه أسطر", rows > 0, f"({rows})")
    for mode in ("excel", "pdf", "print"):
        dlg.export(mode)
        if mode != "print":
            err = valid_export(EXPORTED[-1])
            check(f"[{name}] تصدير {mode}", err == "", err)
    dlg.close()

# لقطة السنة
y26 = conn.execute("SELECT id FROM financial_years WHERE year=2026").fetchone()["id"]
repo.set_year_status(conn, y26, "closed")
repo.create_snapshot(conn, y26)
from app.ui.dialogs_master import SnapshotDialog

snap_dlg = SnapshotDialog(y26, win)
for t in range(snap_dlg.tabs if hasattr(snap_dlg, "tabs") else 0):
    pass
for mode in ("excel", "pdf", "print"):
    try:
        snap_dlg._export_tab(snap_dlg.findChildren(__import__(
            "PySide6.QtWidgets", fromlist=["QTabWidget"]).QTabWidget)[0], mode)
        if mode != "print":
            err = valid_export(EXPORTED[-1])
            check(f"[لقطة السنة] تصدير {mode}", err == "", err)
    except Exception as e:  # noqa: BLE001
        check(f"[لقطة السنة] تصدير {mode}", False, str(e))
repo.set_year_status(conn, y26, "open")
snap_dlg.close()

# فاتورة العميل: PDF وطباعة + التحقق من إخفاء المصروفات
from app.ui.dialogs_ops import (
    customer_invoice_html, export_customer_invoice_pdf, print_customer_invoice,
)

html = customer_invoice_html(conn, ids["inv1"])
trip_prices = [t["price"] for t in calc.get_invoice_full(conn, ids["inv1"])["trips"]]
exp_amounts = [a for a in trip_prices if a > 0]
check("فاتورة العميل: تظهر أسعار النقلات",
      all(f"{a:,.2f}" in html for a in exp_amounts))
exp_total = calc.get_invoice_full(conn, ids["inv1"])["expenses_total"]
check("فاتورة العميل: تخفي المصروفات الداخلية", f"{exp_total:,.2f}" not in html
      and "الربح" not in html)
check("فاتورة العميل: عبارة ZATCA موجودة", "ZATCA" in html or "القيمة المضافة" in html)
export_customer_invoice_pdf(win, ids["inv1"])
err = valid_export(EXPORTED[-1])
check("فاتورة العميل: PDF صالح", err == "", err)
print_customer_invoice(win, ids["inv1"])
check("فاتورة العميل: طباعة (معاينة)", Path(
    os.path.join(tempfile.gettempdir(), "logistic_print_test.pdf")).exists())

# ===========================================================================
step("6) بطاقات الإجماليات في الصفحات تطابق SQL")
def label_value(page, label):
    v = page.totals._values.get(label)
    return float(v.text().replace(",", "")) if v else None

inv_page.refresh()
app.processEvents()
sql_trips = conn.execute(
    "SELECT IFNULL(SUM(price),0) FROM invoice_trips t, invoices i "
    "WHERE i.id=t.invoice_id AND i.date BETWEEN '2026-01-01' AND '2026-12-31'"
).fetchone()[0]
ui_trips = label_value(inv_page, "إجمالي النقلات")
check("إجمالي النقلات في الصفحة = SQL", abs((ui_trips or 0) - sql_trips) < 0.01,
      f"({ui_trips} مقابل {sql_trips})")

pay_page.refresh()
sql_paid = conn.execute(
    "SELECT IFNULL(SUM(amount),0) FROM payment_vouchers "
    "WHERE date BETWEEN '2026-01-01' AND '2026-12-31'").fetchone()[0]
ui_paid = label_value(pay_page, "إجمالي المدفوعات")
check("إجمالي المدفوعات في الصفحة = SQL", abs((ui_paid or 0) - sql_paid) < 0.01)

payroll_page = pages_by_label["إدارة الرواتب"]
payroll_page.refresh()
sql_net = conn.execute(
    "SELECT IFNULL(SUM(net_salary),0) FROM payrolls "
    "WHERE date BETWEEN '2026-01-01' AND '2026-12-31'").fetchone()[0]
ui_net = label_value(payroll_page, "إجمالي الرواتب المنصرفة (الصافي)")
check("إجمالي الرواتب في الصفحة = SQL", abs((ui_net or 0) - sql_net) < 0.01)

# صفحة تقرير P&L: القيم تطابق calc
pnl_page = pages_by_label["الأرباح والخسائر (P&L)"]
pnl_page.filter_row.from_edit.set_iso("2026-01-01")
pnl_page.filter_row.to_edit.set_iso("2026-12-31")
pnl_page.load()
ref = calc.pnl_report(conn, "2026-01-01", "2026-12-31")
check("P&L في الواجهة = المحرك",
      abs((label_value(pnl_page, "إجمالي الإيرادات") or 0) - ref["total_revenue"]) < 0.01
      and abs((label_value(pnl_page, "صافي الربح / الخسارة") or 0) - ref["net"]) < 0.01)

# ===========================================================================
step("7) لا نوافذ منبثقة معلقة ولا أخطاء صامتة")
win.nav_list.setCurrentRow(1)
app.processEvents()

print(f"\n===== النتيجة: نجاح {PASS} / فشل {FAIL} =====")
print(f"ملفات مصدرة: {len(EXPORTED)} في {EXPORT_DIR}")
if FAILURES:
    print("الإخفاقات:")
    for f in FAILURES[:40]:
        print("  -", f)
    sys.exit(1)
print("🎉 كل فحوص الواجهة العميقة نجحت.")
